#  Copyright (c) 2023.
#  Damian TBP

import os
import Boto3Connect
import datetime
import chardet
from openpyxl import Workbook, load_workbook


class EjecucionScriptAutomatizadaAWSAthena:
    def __init__(self, log_dir='.', sql_dir='.', create_excel=False):
        self.log_file = self._generate_log_file_name(log_dir)
        self.sql_dir = sql_dir
        self.create_excel = create_excel
        self.processed_files = []

    def _generate_log_file_name(self, log_dir):
        """
        Genera el nombre del archivo de log basado en la fecha actual.

        Args:
            log_dir (str): Ruta del directorio donde se guardará el archivo de log.

        Returns:
            str: Nombre del archivo de log.
        """
        today = datetime.date.today().strftime("%Y-%m-%d")
        log_file = f"log-{today}.txt"
        return os.path.join(log_dir, log_file)

    def _create_or_open_excel_file(self):
        excel_file = f"log-{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"

        if os.path.exists(excel_file):
            workbook = load_workbook(excel_file)
        else:
            workbook = Workbook()

        return workbook, excel_file

    def _write_to_excel(self, workbook, archivo, inicio, fin, estado):
        sheet = workbook.active

        # Verificar si se necesita crear el encabezado
        if sheet.cell(row=1, column=1).value is None:
            sheet.cell(row=1, column=1).value = "Nombre de archivo SQL"
            sheet.cell(row=1, column=2).value = "Fecha y hora de inicio de ejecución"
            sheet.cell(row=1, column=3).value = "Fecha y hora de fin de ejecución"
            sheet.cell(row=1, column=4).value = "Estado"

        if archivo not in self.processed_files:
            row_count = sheet.max_row + 1
            sheet.cell(row=row_count, column=1).value = archivo
            self.processed_files.append(archivo)
        else:
            row_count = self._get_row_for_file(sheet, archivo)

        sheet.cell(row=row_count, column=2).value = inicio
        sheet.cell(row=row_count, column=3).value = fin
        sheet.cell(row=row_count, column=4).value = estado

    def _get_row_for_file(self, sheet, archivo):
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == archivo:
                return row
        return sheet.max_row + 1

    def buscar_archivos_sql(self):
        """
        Busca los archivos con extensión .sql en el directorio actual y los devuelve en orden alfabético.

        Returns:
            list: Lista de archivos .sql encontrados en el directorio actual.
        """
        archivos_sql = []
        for file in sorted(os.listdir(self.sql_dir)):
            if file.endswith('.sql'):
                archivos_sql.append(file)
        return archivos_sql

    @staticmethod
    def _read_file_content(file_path):
        with open(file_path, 'rb') as f:
            rawdata = f.read()
            result = chardet.detect(rawdata)
            encoding = result['encoding']
            return rawdata.decode(encoding)

    def obtener_queries(self, archivo):
        """
        Obtiene las sentencias SQL de un archivo.

        Args:
            archivo (str): Ruta y nombre del archivo.

        Returns:
            list: Lista de sentencias SQL encontradas en el archivo.
        """
        queries = []
        file_content = self._read_file_content(os.path.join(self.sql_dir, archivo))
        contenido = file_content.strip()
        inicio = '--<sql>--'
        fin = '--</sql>--'
        while inicio in contenido and fin in contenido:
            inicio_pos = contenido.index(inicio)
            fin_pos = contenido.index(fin)
            query = contenido[inicio_pos + len(inicio):fin_pos].strip()
            queries.append(query)
            contenido = contenido[fin_pos + len(fin):]
        return queries

    def ejecutar_queries(self, queries, query_error, archivo, workbook=None):
        """
        Ejecuta las sentencias SQL y registra los resultados en el archivo de log y/o Excel.

        Args:
            queries (list): Lista de sentencias SQL a ejecutar.
            query_error (int): Índice de la última consulta que dio error en la ejecución anterior.
            archivo (str): Nombre del archivo que contiene las sentencias SQL.
            workbook (Workbook, optional): Objeto Workbook de openpyxl para escribir en el archivo de Excel.
                                          None si no se desea escribir en Excel. Por defecto es None.

        Returns:
            bool: Indica si todas las consultas se ejecutaron correctamente o si se detuvo la ejecución debido a un error.
        """
        with open(self.log_file, 'a') as log:
            for query in queries[query_error:]:
                try:
                    # Conexión a Boto3 utilizando la clase Boto3Connect
                    # El parámetro log=True indica que se deben registrar los mensajes de ejecución
                    boto3_instance = Boto3Connect.Boto3Connect(log=True)
                    boto3_instance.run_query(query)
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    log_message = f"\n[{timestamp}] (Archivo: {archivo})\nQuery ejecutado con ÉXITO: {query[:120]}...\n"
                    log.write(log_message)
                    if workbook:
                        self._write_to_excel(workbook, archivo, timestamp, timestamp, "ÉXITO")
                except Exception as e:
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    log_message = f"\n[{timestamp}] (Archivo: {archivo})\nERROR al ejecutar el query: {query}\n"
                    log_message += f"[{timestamp}] Detalles del error: {str(e)}\n"
                    log_message += f"[{timestamp}] Deteniendo la ejecución de las queries restantes.\n"
                    log.write(log_message)
                    if workbook:
                        self._write_to_excel(workbook, archivo, timestamp, timestamp, "ERROR")
                    return False
        return True

    import os

    def ejecutar_script(self, archivo_inicial=None, archivo_final=None):
        """
        Ejecuta el script principal.

        Realiza la lectura de los archivos .sql en orden alfabético,
        obtiene las sentencias SQL de cada archivo y las ejecuta utilizando Boto3Connect.
        Registra los resultados de la ejecución en el archivo de log y/o Excel.

        Args:
            archivo_inicial (str, optional): Nombre del archivo SQL desde el cual se debe iniciar la ejecución.
                                            None para comenzar desde el primer archivo. Por defecto es None.
            archivo_final (str, optional): Nombre del último archivo SQL que debe contener la lista "archivos_sql".
                                           None para procesar todos los archivos. Por defecto es None.
        """
        archivos_sql = self.buscar_archivos_sql()
        query_error = 0

        # Agregar encabezado de nueva ejecución
        timestamp_inicio = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(self.log_file, 'a') as log:
            log.write(f"[{timestamp_inicio}] -------- Inicio de ejecución --------\n")
            print(f"[{timestamp_inicio}] -------- Inicio de ejecución --------\n")

        workbook = None
        if self.create_excel:
            workbook, excel_file = self._create_or_open_excel_file()

        ejecutar_archivo = False
        for archivo in archivos_sql:
            if archivo_inicial is None or archivo == archivo_inicial:
                ejecutar_archivo = True
            if ejecutar_archivo:
                print(f"Procesando Archivo: {archivo}\n")
                queries = self.obtener_queries(archivo)
                if query_error > 0:
                    queries = queries[query_error:]
                    query_error = 0
                if not self.ejecutar_queries(queries, query_error, archivo, workbook):
                    break
                else:
                    query_error = 0

            if archivo == archivo_final:
                break

        # Guardar archivo de Excel si se creó
        if workbook:
            workbook.save(excel_file)

        # Agregar finalización de la ejecución
        timestamp_fin = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(self.log_file, 'a') as log:
            log.write(f"\n[{timestamp_fin}] -------- Fin de ejecución --------\n\n")
            print(f"\n[{timestamp_fin}] -------- Fin de ejecución --------\n\n")


# Ejemplo de uso
if __name__ == '__main__':
    sql_dir = "C:\\Users\\dpara\\OneDrive\\Documentos\\git\\typ\\scripts\\Entregable 18.08.2023\\"
    ejecucion_script = EjecucionScriptAutomatizadaAWSAthena(log_dir='.', sql_dir=sql_dir, create_excel=True)
    ejecucion_script.ejecutar_script(archivo_inicial='2023.08.18 step 43 - consume experiencia_laboral.sql', archivo_final='2023.08.18 step 45 - consume organizacion_actividad.sql')
    # ejecucion_script.ejecutar_script()